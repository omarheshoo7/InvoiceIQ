import io

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Colour palette ────────────────────────────────────────────────────────────
DARK_BLUE  = "2E4057"
LIGHT_GREY = "F0F4F8"
MID_GREY   = "CBD5E0"
WHITE      = "FFFFFF"


def export_to_excel(fields: dict, line_items_df: pd.DataFrame) -> bytes:
    """
    Build a styled Excel workbook from invoice fields + line items.
    Returns raw bytes ready for st.download_button.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice"

    _write_title(ws)
    _write_fields(ws, fields)
    _write_line_items(ws, line_items_df)
    _auto_fit_columns(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── Section writers ───────────────────────────────────────────────────────────

def _write_title(ws):
    ws.merge_cells("A1:D1")
    cell = ws["A1"]
    cell.value = "INVOICE"
    cell.font = Font(bold=True, size=18, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32


def _write_fields(ws, fields: dict):
    labels = {
        "invoice_number": "Invoice Number",
        "invoice_date":   "Invoice Date",
        "vendor_name":    "Vendor / Company",
        "total_amount":   "Total Amount",
        "currency":       "Currency",
    }
    label_font = Font(bold=True)
    row_fill   = PatternFill("solid", fgColor=LIGHT_GREY)

    for i, (key, label) in enumerate(labels.items(), start=3):
        label_cell = ws.cell(row=i, column=1, value=label)
        label_cell.font  = label_font
        label_cell.fill  = row_fill

        value_cell = ws.cell(row=i, column=2, value=fields.get(key, ""))
        value_cell.fill  = row_fill


def _write_line_items(ws, df: pd.DataFrame):
    START = 10  # Leave a gap below the field section

    # Section heading
    ws.merge_cells(f"A{START}:D{START}")
    heading = ws[f"A{START}"]
    heading.value     = "LINE ITEMS"
    heading.font      = Font(bold=True, color=WHITE)
    heading.fill      = PatternFill("solid", fgColor=DARK_BLUE)
    heading.alignment = Alignment(horizontal="center")

    if df is None or df.empty:
        ws.cell(row=START + 1, column=1, value="No line items recorded.")
        return

    thin = Side(style="thin")
    border = Border(bottom=thin, right=thin)

    # Column headers
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=START + 1, column=col_idx, value=col_name)
        cell.font      = Font(bold=True)
        cell.fill      = PatternFill("solid", fgColor=MID_GREY)
        cell.border    = border
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_offset, (_, row) in enumerate(df.iterrows(), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=START + row_offset, column=col_idx, value=value)
            cell.border = border


def _auto_fit_columns(ws):
    for col in ws.columns:
        letter   = get_column_letter(col[0].column)
        max_len  = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[letter].width = min(max_len + 4, 50)
